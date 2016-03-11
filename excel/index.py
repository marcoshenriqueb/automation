import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
import os

wb = openpyxl.load_workbook(os.path.join(os.getcwd(), 'example.xlsx'))
# print(wb.get_sheet_names())
# print(wb.active.title)
sheet = wb.get_sheet_by_name(wb.active.title)
# print(sheet["A1"].value)
# c = sheet["B1"]
# print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
# print('Cell ' + c.coordinate + ' is ' + c.value)
# print(sheet.cell(row=2, column=2).value)
# print('The last line is ' + str(sheet.max_row))
# print('The last column is ' + str(sheet.max_column))
# print(get_column_letter(30))
# print(column_index_from_string('AD'))

# print(tuple(sheet['A1':'C3']))
# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')

# print(sheet.columns[1])
# for cellObj in sheet.columns[1]:
#     print(cellObj.value)
