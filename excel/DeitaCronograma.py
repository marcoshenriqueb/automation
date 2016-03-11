import openpyxl
import warnings
import os


class DeitaCronograma():
    """DeitaCronograma prepara a planilha para o planejamento fim da obra"""
    def __init__(self, workbook):
        warnings.filterwarnings("ignore")
        self.workbook = workbook
        self.titleRow = None
        self.lastRow = None
        self.wb = openpyxl.load_workbook(os.path.join(os.getcwd(), workbook + '.xlsx'))

    def __getTitleRow(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        col = sheet.columns[0]
        for cellObj in col:
            if cellObj.value == 'Descrição':
                self.titleRow = cellObj.row
                break

    def __getLastRow(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        row = self.titleRow + 1
        while sheet.cell(row=row, column=1).value != None:
            if isinstance(sheet.cell(row=row, column=6).value, int) or \
               (type(sheet.cell(row=row, column=6).value) == str and \
               sheet.cell(row=row, column=6).value[0] == "="):
                self.lastRow = row

            row += 1

    def __writeTitles(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        sheet.cell(row=self.titleRow, column=7).value = 'Previsão de Gastos'
        sheet.cell(row=self.titleRow, column=8).value = 'Valor Fechado'
        sheet.cell(row=self.titleRow, column=9).value = 'Valores Pagos'
        sheet.cell(row=self.titleRow, column=10).value = 'Saldo a pagar'

    def __createPagamentosRealizadosSpreasheet(self):
        ws = self.wb.create_sheet()
        ws.title = 'Pagamentos realizados'
        ws['A1'].value = 'Índice'
        ws['B1'].value = 'Descrição'
        ws['C1'].value = 'Valor Pago'
        ws['D1'].value = 'Data de Pagamento'
        ws['E1'].value = 'N de Parcelas'
        ws['F1'].value = 'Intervalo'
        ws['G1'].value = 'Valor Total'

    def __writePrevisaoDeGastosFormula(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        for row in range(self.titleRow + 1, self.lastRow + 1):
            col6Val = sheet.cell(row=row, column=6).value
            if isinstance(col6Val, int) or (type(col6Val) == str and col6Val[0] == "="):
                sheet.cell(row=row, column=7).value = '=F' + str(row) + '*0.7'
            row += 1

    def __writeSaldoAPagarFormula(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        row = self.titleRow + 1
        for row in range(self.titleRow + 1, self.lastRow + 1):
            col6Val = sheet.cell(row=row, column=6).value
            if isinstance(col6Val, int) or (type(col6Val) == str and col6Val[0] == "="):
                sheet.cell(row=row, column=10).value = '=IF(ISNUMBER(H' + str(row) + '-I' + str(row) + '),\
                                                            H' + str(row) + '-I' + str(row) + ',\
                                                            "")'
            row += 1

    def __writeDescricaoFormula(self):
        sheet = self.wb.get_sheet_by_name('Pagamentos realizados')
        firstRow = self.titleRow + 1
        lastRow = self.lastRow
        for row in range(2, 500):
            formula = '=IF(\
                        ISERROR(\
                        VLOOKUP($A' + str(row) + ',\
                        \'PLANILHA BASE\'!$A$' + str(firstRow) + ':$F$' + str(lastRow) + ',2,0)),\
                        "-",\
                        VLOOKUP($A' + str(row) + ',\
                        \'PLANILHA BASE\'!$A$' + str(firstRow) + ':$F$' + str(lastRow) + ',2,0))'
            sheet.cell(row=row, column=2).value = formula

    def __writeValoresPagosFormula(self):
        sheet = self.wb.get_sheet_by_name('PLANILHA BASE')
        for row in range(self.titleRow + 1, self.lastRow + 1):
            col1Val = sheet.cell(row=row, column=1).value
            col6Val = sheet.cell(row=row, column=6).value
            if type(col6Val) == int:
                formula = '=SUMIF(\'Pagamentos realizados\'!$A$2:$A$500,\
                                  \'PLANILHA BASE\'!$A' + str(row) + ',\
                                  \'Pagamentos realizados\'!$C$2:$C$500)'
                sheet.cell(row=row, column=9).value = formula
            elif type(col6Val) == str and col6Val[0] == "=":
                formula = col6Val.replace('F', 'I')
                sheet.cell(row=row, column=9).value = formula

    def transform(self):
        self.__getTitleRow()
        self.__getLastRow()
        self.__writeTitles()
        self.__writePrevisaoDeGastosFormula()
        self.__writeSaldoAPagarFormula()
        self.__createPagamentosRealizadosSpreasheet()
        self.__writeDescricaoFormula()
        self.__writeValoresPagosFormula()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.wb.save(filename = self.workbook + '_m.xlsx')
