import openpyxl
from openpyxl.styles import Font, PatternFill, colors, Alignment
import warnings
import os


class FormataCronograma():
    """FromataCronograma formata a planilha para o planejamento fim da obra"""
    def __init__(self, workbook):
        warnings.filterwarnings("ignore")
        self.workbook = workbook
        self.titleRow = None
        self.lastRow = None
        self.wb = openpyxl.load_workbook(os.path.join(os.getcwd(), workbook + '_m.xlsx'))

    def style(self):
        self.__stylePagamentosRealizados()
        self.__stylePlanilhaBase()


    def __stylePagamentosRealizados(self):
        ws = self.wb.get_sheet_by_name('Pagamentos realizados')
        row = ws.rows[0]
        for cell in row:
            cell.font = Font(color=colors.WHITE,
                             bold=True)
            cell.fill = PatternFill(start_color=colors.DARKBLUE,
                                   end_color=colors.DARKBLUE,
                                   fill_type='solid')

        for row in ws['A1':'G500']:
            for cell in row:
                cell.alignment = Alignment(horizontal='center')


        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15

    def __stylePlanilhaBase(self):
        ws = self.wb.get_sheet_by_name('PLANILHA BASE')

        for row in ws['G1':'J200']:
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15


    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.wb.save(filename = self.workbook + '_m.xlsx')
